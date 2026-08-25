# -*- coding: utf-8 -*-
import os
import sys
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
import docx
from docx.shared import Inches, Pt, RGBColor
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.enum.table import WD_TABLE_ALIGNMENT
from docx.oxml import parse_xml
from docx.oxml.ns import nsdecls

OUTPUT_S2_DIR = "/Users/ludwin/Documents/Universidad/2026-2/ingenieria-de-software/2_Tareas/sprint-2"
OUTPUT_TAREAS_DIR = "/Users/ludwin/Documents/Universidad/2026-2/ingenieria-de-software/2_Tareas"
WORKSPACE_DEV = "/Users/ludwin/Developer/ironlink_workspace"
WORKSPACE_DOCS = "/Users/ludwin/Documents/Universidad/2026-2/ingenieria-de-software/4_Proyectos_y_Examenes/ironlink_workspace"

SCREENSHOTS_DIR = os.path.join(WORKSPACE_DEV, "tests/e2e/screenshots_desktop")
DIAGRAMS_DIR = os.path.join(WORKSPACE_DEV, "tests/e2e/diagrams")

os.makedirs(OUTPUT_S2_DIR, exist_ok=True)
os.makedirs(OUTPUT_TAREAS_DIR, exist_ok=True)

# ─── TEAM DATA ─────────────────────────────────────────────────────────────
TEAM_MEMBERS = [
    ("Ludwin Saul Vasquez Romero", "Scrum Master / Backend & Architecture Lead"),
    ("Luis Alexander Rivera Alvarez", "QA Lead / Database & Security Dev"),
    ("Alberto Jose Velazquez Paz", "Frontend Lead / Desktop UI & QA Tester"),
    ("Luis Angel Zuniga Menjivar", "Backend Dev / API Security & Conformance"),
    ("Ricardo Alberto Mendiola Hernandez", "Dev / Chat Persistente & Perfil Lead"),
    ("Victor Arnoldo Iglesias Sandoval", "Dev / Reuniones & Servicios Sincronos"),
    ("Jose Luis Fuentes Ochoa", "Dev / Subgrupos & Organizacion de Nodos")
]

TEAM_MEMBERS_TEXT = [
    "Ludwin Saul Vasquez Romero (Scrum Master / Backend & Architecture Lead)",
    "Luis Alexander Rivera Alvarez (QA Lead / Database & Security Dev)",
    "Alberto Jose Velazquez Paz (Frontend Lead / Desktop UI & QA Tester)",
    "Luis Angel Zuniga Menjivar (Backend Dev / API Security & Conformance)",
    "Ricardo Alberto Mendiola Hernandez (Dev / Chat Persistente & Perfil Lead)",
    "Victor Arnoldo Iglesias Sandoval (Dev / Reuniones & Servicios Sincronos)",
    "Jose Luis Fuentes Ochoa (Dev / Subgrupos & Organizacion de Nodos)"
]

# ─── TEST CASES DATA (23 CASOS DE PRUEBA SPRINT 2) ─────────────────────────
ALL_TEST_CASES_S2 = [
    # MÓDULO 1: CHAT PERSISTENTE EN CANALES (IRL-WKS-US-03)
    ("TC-CHT-001", "Chat en Vivo", "Ricardo Mendiola", "Envio y persistencia de mensaje en canal con usuario activo", "IRL-WKS-US-03",
     "Dado que el usuario Tester QA esta en el chat del nodo\nCuando escribe un mensaje y presiona Enviar\nEntonces se inserta en PostgreSQL y se renderiza en pantalla con su avatar y rol",
     "Usuario autenticado en la aplicacion de escritorio dentro de la vista de chat del nodo.",
     "1. Abrir la aplicacion de escritorio e iniciar sesion como Tester QA.\n2. Acceder al espacio de trabajo del Nodo colaborativo.\n3. Seleccionar la pestaña de [Chat] en la cabecera.\n4. Escribir el mensaje: Hola equipo InnovaSoft, probando chat persistente de Sprint 2!\n5. Presionar el boton Enviar o la tecla Enter.",
     "El mensaje se envia mediante POST /nodos/{id}/mensajes, se almacena en PostgreSQL y se renderiza en la pantalla con el avatar, nombre Tester QA, rol y timestamp actual.",
     "El mensaje fue enviado y persistido exitosamente en 8 ms. Aparece en pantalla con el formato corporativo y queda registrado en la base de datos.",
     "Alta", "Funcional", "Aprobado", "Pasa", "Alberto Velazquez", "2h", "09_nodo_chat_message_sent.png"),
     
    ("TC-CHT-002", "Chat en Vivo", "Ricardo Mendiola", "Carga historica de chat cronologica y auto-scroll inteligente", "IRL-WKS-US-03",
     "Dado que existen mensajes previos guardados en la tabla mensajes\nCuando el usuario entra al chat\nEntonces carga los mensajes en orden created_at ASC y realiza auto-scroll al final",
     "Existen mensajes previos guardados en la tabla mensajes para el nodo seleccionado.",
     "1. Abrir el canal de chat del nodo.\n2. Observar la carga inicial de los mensajes.\n3. Verificar el orden cronologico (created_at ASC) y la posicion del scroll.",
     "La lista de mensajes carga de forma inmediata y el ScrollController se desplaza suavemente hacia el ultimo mensaje recibido en la parte inferior.",
     "Carga historica completa en 6 ms. El scroll automatico funciono de forma reactiva sin desbordamiento de componentes.",
     "Media", "Interfaz", "Aprobado", "Pasa", "Alberto Velazquez", "1.5h", "08_nodo_chat_workspace.png"),
     
    ("TC-CHT-003", "Chat en Vivo", "Ricardo Mendiola", "Identificacion visual del autor (Avatar, Nombre y Rol) en burbujas", "IRL-WKS-US-03",
     "Dado que se renderizan mensajes en el canal\nCuando se visualiza cada burbuja\nEntonces muestra el circulo de avatar con color asignado, nombre del autor y etiqueta de rol",
     "Mensajes enviados por usuarios con diferentes roles (OWNER, ADMIN, MEMBER).",
     "1. Visualizar la lista de mensajes en el chat.\n2. Comprobar la presencia del circulo de avatar con su color hexadecimal.\n3. Validar el nombre del autor y la insignia de rol.",
     "Cada mensaje muestra claramente la identidad del remitente, respetando la paleta de colores y el rol asignado en el nodo.",
     "Identificacion visual verificada al 100%. Formato limpio y consistente con la linea grafica.",
     "Media", "UI / UX", "Aprobado", "Pasa", "Alberto Velazquez", "1.5h", "09_nodo_chat_message_sent.png"),
     
    ("TC-CHT-004", "Chat en Vivo", "Ricardo Mendiola", "Validacion de mensaje vacio o solo espacios en blanco", "IRL-WKS-US-03",
     "Dado que el campo de texto esta vacio o contiene solo espacios\nCuando el usuario presiona Enviar\nEntonces el boton permanece inhabilitado o no realiza ninguna peticion al backend",
     "Usuario en el canal de chat con el campo de texto en blanco.",
     "1. Dejar el campo de texto vacio.\n2. Intentar enviar presionando la tecla Enter.\n3. Escribir espacios en blanco y presionar Enviar.",
     "La aplicacion valida que el texto no este vacio antes de emitir la peticion HTTP, evitando inserciones innecesarias en la base de datos.",
     "Validacion de texto en blanco exitosa. No se registraron peticiones vacias en el servidor.",
     "Baja", "Validacion", "Aprobado", "Pasa", "Luis Zuniga", "1h", "08_nodo_chat_workspace.png"),
     
    ("TC-CHT-005", "Chat en Vivo", "Ricardo Mendiola", "Bloqueo de acceso al chat a usuarios no miembros (403 Forbidden)", "IRL-WKS-US-03",
     "Dado un usuario que no pertenece al nodo\nCuando intenta consultar o enviar mensajes al endpoint /nodos/{id}/mensajes\nEntonces el servidor Rust deniega el acceso con HTTP 403 Forbidden",
     "Usuario autenticado pero sin registro en nodo_miembros para el nodo objetivo.",
     "1. Enviar peticion GET /nodos/{id_nodo_ajeno}/mensajes con token de usuario no miembro.\n2. Medir tiempo de respuesta y codigo HTTP.",
     "El backend valida la membresia en la tabla nodo_miembros y rechaza inmediatamente con 403 Forbidden.",
     "Acceso bloqueado en 4 ms con codigo HTTP 403 Forbidden. Seguridad Fail-Closed verificada.",
     "Alta", "Seguridad / RBAC", "Aprobado", "Pasa", "Luis Rivera", "2h", "diag_06_chat_messaging.png"),

    # MÓDULO 2: SUBGRUPOS DE NODO (IRL-WKS-US-02)
    ("TC-SUB-001", "Subgrupos", "Jose Fuentes", "Creacion exitosa de subgrupo publico con auto-asignacion", "IRL-WKS-US-02",
     "Dado que el usuario es miembro del nodo\nCuando ingresa nombre y descripcion en Nuevo Subgrupo\nEntonces crea el subgrupo, auto-asocia al creador y lo lista con 1 miembro",
     "Usuario con sesion activa y miembro del nodo en la pestaña de Subgrupos.",
     "1. Hacer clic en la pestaña [Subgrupos] en la barra de navegacion del nodo.\n2. Presionar el boton Nuevo Subgrupo.\n3. Ingresar el nombre Frontend & UI y descripcion Celula de trabajo de interfaz.\n4. Dejar el switch de privacidad en Subgrupo Publico.\n5. Presionar Crear Subgrupo.",
     "El sistema crea el subgrupo en la tabla subgrupos, asocia automaticamente al creador en subgrupo_miembros y lo muestra en la lista con contador de 1 miembro.",
     "Subgrupo creado exitosamente en 14 ms. Se renderiza la tarjeta en la cuadricula de subgrupos con su nombre e icono de grupo publico.",
     "Alta", "Funcional / DB", "Aprobado", "Pasa", "Luis Rivera", "2h", "s2_02_subgrupos_view.png"),
     
    ("TC-SUB-002", "Subgrupos", "Jose Fuentes", "Creacion de subgrupo privado y aislamiento de visibilidad", "IRL-WKS-US-02",
     "Dado que el usuario activa el switch Subgrupo Privado\nCuando guarda el subgrupo\nEntonces se registra con flag es_privado=true y badge Privado con candado",
     "Usuario en el dialogo modal de Nuevo Subgrupo.",
     "1. Abrir modal Nuevo Subgrupo.\n2. Ingresar nombre Ciberseguridad & Kernel.\n3. Activar el switch Subgrupo Privado.\n4. Presionar Crear Subgrupo.",
     "El subgrupo se registra con flag es_privado=true, mostrando un candado e insignia Privado, restringiendo el acceso unicamente a invitados.",
     "Subgrupo privado registrado correctamente. La interfaz muestra el candado cian y la etiqueta Privado.",
     "Media", "Seguridad / Logica", "Aprobado", "Pasa", "Alberto Velazquez", "1.5h", "s2_03_create_subgrupo_dialog.png"),
     
    ("TC-SUB-003", "Subgrupos", "Jose Fuentes", "Validacion de nombre obligatorio y longitud en creacion de subgrupo", "IRL-WKS-US-02",
     "Dado que el usuario intenta crear un subgrupo con nombre vacio\nCuando presiona Crear Subgrupo\nEntonces el formulario muestra advertencia en rojo y bloquea el envio",
     "Modal de creacion de subgrupo abierto.",
     "1. Dejar el campo de nombre vacio.\n2. Presionar el boton Crear Subgrupo.\n3. Verificar mensaje de error visual.",
     "El campo de texto se bordea en rojo y muestra el mensaje El nombre del subgrupo es obligatorio.",
     "Validacion visual y de negocio completada exitosamente. No se enviaron peticiones invalidas.",
     "Media", "Validacion UI", "Aprobado", "Pasa", "Alberto Velazquez", "1h", "s2_03_create_subgrupo_dialog.png"),
     
    ("TC-SUB-004", "Subgrupos", "Jose Fuentes", "Ciclo dinamico de membresia: Unirse a subgrupo (Join)", "IRL-WKS-US-02",
     "Dado un subgrupo publico existente en el nodo\nCuando el usuario presiona Unirse\nEntonces se inserta en subgrupo_miembros y el contador de integrantes incrementa",
     "Subgrupo publico creado y visible en la lista.",
     "1. Localizar un subgrupo en la lista donde el usuario no sea miembro.\n2. Presionar el boton Unirse.\n3. Verificar que el boton cambie de estado y el contador aumente a 2 miembros.",
     "Peticion POST /nodos/{id}/subgrupos/{subgrupo_id}/join exitosa, registrando la membresia en la base de datos.",
     "Union a subgrupo completada en 10 ms. Interfaz actualizada reactivamente.",
     "Alta", "Integracion / ACID", "Aprobado", "Pasa", "Luis Zuniga", "1.5h", "s2_02_subgrupos_view.png"),
     
    ("TC-SUB-005", "Subgrupos", "Jose Fuentes", "Ciclo dinamico de membresia: Salir de subgrupo (Leave)", "IRL-WKS-US-02",
     "Dado un usuario miembro de un subgrupo\nCuando presiona Salir\nEntonces se elimina de subgrupo_miembros y el contador decrementa",
     "Usuario con membresia activa en un subgrupo.",
     "1. En la tarjeta del subgrupo, presionar el boton Salir.\n2. Confirmar la accion en el dialogo de confirmacion.\n3. Comprobar la eliminacion del registro en la base de datos.",
     "Peticion POST .../leave ejecutada, eliminando el registro en subgrupo_miembros y decrementando el contador.",
     "Salida de subgrupo exitosa en 9 ms. Contador actualizado de forma atomica.",
     "Alta", "Integracion / ACID", "Aprobado", "Pasa", "Luis Zuniga", "1.5h", "s2_02_subgrupos_view.png"),
     
    ("TC-SUB-006", "Subgrupos", "Jose Fuentes", "Eliminacion de subgrupo por creador/admin y cascada de datos", "IRL-WKS-US-02",
     "Dado que el creador del subgrupo o un OWNER/ADMIN solicita su eliminacion\nCuando confirma la accion\nEntonces se elimina de subgrupos y se purgan sus miembros en cascada",
     "Subgrupo creado con miembros asociados.",
     "1. Iniciar sesion como creador del subgrupo o Admin del nodo.\n2. Presionar el icono de eliminar en la tarjeta del subgrupo.\n3. Confirmar la eliminacion.\n4. Validar en PostgreSQL que no queden registros huerfanos.",
     "El subgrupo se elimina de la base de datos y la clausula ON DELETE CASCADE purga todas las relaciones asociadas.",
     "Eliminacion en cascada ejecutada perfectamente en 12 ms con 0 huerfanos.",
     "Critica", "ACID / Cascada", "Aprobado", "Pasa", "Ludwin Romero", "2h", "diag_08_subgrupos.png"),

    # MÓDULO 3: CALENDARIO Y REUNIONES SÍNCRONAS (IRL-WKS-US-04)
    ("TC-REU-001", "Reuniones", "Victor Iglesias", "Programacion de sesion con timestamps ISO 8601 UTC y Meet", "IRL-WKS-US-04",
     "Dado que el usuario completa titulo, fecha/hora, duracion y link Google Meet\nCuando presiona Programar Sesion\nEntonces se guarda en PostgreSQL en UTC y se visualiza en la agenda",
     "Usuario miembro del nodo en la pestaña de Reuniones.",
     "1. Hacer clic en la pestaña [Reuniones] en la cabecera del nodo.\n2. Presionar el boton Programar Sesion.\n3. Completar titulo (Daily Scrum InnovaSoft), fecha y hora futura.\n4. Seleccionar duracion de 30 min.\n5. Ingresar enlace https://meet.google.com/abc-defg-hij y presionar Programar Sesion.",
     "Se inserta la reunion en PostgreSQL con timestamp ISO 8601 UTC y se muestra en la agenda con tarjeta detallada y boton Unirse a Meet.",
     "Reunion guardada exitosamente en 11 ms. Tarjeta renderizada en la agenda con fecha formateada e insignia ● Programada.",
     "Alta", "Protocolos / Negocio", "Aprobado", "Pasa", "Ludwin Romero", "2h", "s2_04_reuniones_view.png"),
     
    ("TC-REU-002", "Reuniones", "Victor Iglesias", "Selector interactivo de duracion estimada (15, 30, 45, 60, 90 min)", "IRL-WKS-US-04",
     "Dado el dialogo de programacion de reunion\nCuando el usuario hace clic sobre los chips de duracion\nEntonces el chip seleccionado se activa con borde y texto cian",
     "Usuario dentro del modal de programacion de reunion.",
     "1. Abrir dialogo Programar Nueva Reunion.\n2. Probar los chips de duracion (15, 30, 45, 60, 90 min).\n3. Validar el cambio visual de seleccion.",
     "Los chips alternan de estado visual de forma instantanea actualizando el valor de duracion en minutos en el payload.",
     "Selector de duracion validado con exito. Estado reactivo perfecto.",
     "Media", "Interfaz", "Aprobado", "Pasa", "Alberto Velazquez", "1h", "s2_05_create_reunion_dialog.png"),
     
    ("TC-REU-003", "Reuniones", "Victor Iglesias", "Calculo dinamico de insignias de estado (● Programada vs Finalizada)", "IRL-WKS-US-04",
     "Dado que existen reuniones con fechas pasadas y futuras\nCuando se renderizan en el calendario\nEntonces las futuras muestran badge verde ● Programada y las pasadas badge gris",
     "Reuniones existentes en la base de datos con distintas marcas temporales.",
     "1. Abrir la pestaña de [Reuniones].\n2. Observar las insignias de estado de cada tarjeta de reunion.\n3. Comprobar que la reunion futura muestra el punto verde ● Programada.",
     "El componente calcula dinamicamente el estado comparando la fecha de la reunion contra la hora actual del sistema.",
     "Insignias de estado calculadas correctamente sin discrepancias de zona horaria.",
     "Media", "Logica UI", "Aprobado", "Pasa", "Alberto Velazquez", "1.5h", "s2_04_reuniones_view.png"),
     
    ("TC-REU-004", "Reuniones", "Victor Iglesias", "Validacion de URL de videollamada y boton directo Unirse a Meet", "IRL-WKS-US-04",
     "Dado una reunion con enlace a Google Meet\nCuando el usuario presiona Unirse a Meet\nEntonces el cliente invoca el navegador o app de videollamada con la URL exacta",
     "Reunion agendada con URL de Google Meet.",
     "1. Localizar la tarjeta de reunion en el calendario.\n2. Presionar el boton Unirse a Meet.\n3. Validar que la URL se abra sin alteraciones.",
     "El boton activa el lanzador de URLs del sistema abriendo la sala de videollamada configurada.",
     "Enlace verificado y probado exitosamente con apertura instantanea.",
     "Alta", "Integracion", "Aprobado", "Pasa", "Luis Zuniga", "1.5h", "s2_04_reuniones_view.png"),
     
    ("TC-REU-005", "Reuniones", "Victor Iglesias", "Cancelacion y eliminacion de sesion agendada en calendario", "IRL-WKS-US-04",
     "Dado que el organizador o admin decide cancelar una reunion\nCuando presiona el icono de eliminar\nEntonces se elimina de la base de datos y desaparece del calendario",
     "Reunion creada por el usuario autenticado.",
     "1. Presionar el boton de eliminar en la tarjeta de reunion.\n2. Confirmar la cancelacion.\n3. Verificar la actualizacion reactiva de la lista.",
     "Peticion DELETE /nodos/{id}/reuniones/{reunion_id} ejecutada, eliminando el registro en PostgreSQL.",
     "Reunion eliminada en 8 ms. La lista se refresca inmediatamente sin requerir recarga.",
     "Media", "Funcional", "Aprobado", "Pasa", "Luis Rivera", "1.5h", "diag_09_reuniones.png"),

    # MÓDULO 4: IDENTIDAD Y PERFIL DE USUARIO (IRL-IAM-US-05)
    ("TC-PRF-001", "Perfil de Usuario", "Ricardo Mendiola", "Personalizacion de color de avatar entre 8 opciones corporativas", "IRL-IAM-US-05",
     "Dado que el usuario abre el modal de perfil\nCuando selecciona uno de los 8 circulos de color (#00E5FF, #00BFA5, etc.) y guarda\nEntonces actualiza en BD y Riverpod propaga el cambio en toda la UI",
     "Usuario autenticado en la aplicacion de escritorio.",
     "1. Hacer clic en el avatar de usuario en la esquina superior derecha.\n2. Seleccionar color cian (#00E5FF) de la paleta de 8 colores.\n3. Presionar Guardar Cambios.\n4. Observar la actualizacion del avatar en la barra superior y en los mensajes de chat.",
     "Peticion PUT /users/me procesada con exito, almacenando el color en la columna avatar_color y sincronizando el estado global con Riverpod.",
     "Avatar actualizado en 9 ms. El color cian se refleja inmediatamente en toda la app sin parpadeos.",
     "Media", "StateNotifier", "Aprobado", "Pasa", "Alberto Velazquez", "1.5h", "s2_01_profile_dialog.png"),
     
    ("TC-PRF-002", "Perfil de Usuario", "Ricardo Mendiola", "Actualizacion de chip de presencia dinamica y biografia", "IRL-IAM-US-05",
     "Dado que el usuario selecciona un chip de estado (🟢 En linea) y redacta su biografia\nCuando guarda los cambios\nEntonces se actualizan los campos en PostgreSQL y se muestran en su tarjeta de usuario",
     "Usuario en el modal de edicion de perfil.",
     "1. Seleccionar el chip de presencia rapida 🟢 En linea.\n2. Redactar en el campo de biografia: Auditor Lider QA InnovaSoft.\n3. Ingresar numero de telefono +503 7777-8888.\n4. Presionar Guardar Cambios.",
     "Los campos bio, status_text y telefono se guardan en la tabla users y se muestran en el perfil y listas de miembros.",
     "Presencia y biografia actualizadas exitosamente en 9 ms. Datos sincronizados con el backend.",
     "Media", "Funcional UI", "Aprobado", "Pasa", "Alberto Velazquez", "1.5h", "s2_01_profile_dialog.png"),
     
    ("TC-PRF-003", "Perfil de Usuario", "Ricardo Mendiola", "Cambio criptografico de contraseña con verificacion Argon2id", "IRL-IAM-US-05",
     "Dado que el usuario solicita cambio de contraseña\nCuando ingresa clave actual valida y nueva clave con alta entropia\nEntonces genera hash Argon2id con salt de hardware OsRng y actualiza en BD",
     "Usuario autenticado en la seccion de seguridad de su perfil.",
     "1. En el dialogo de perfil, ingresar la contraseña actual valida.\n2. Ingresar la nueva contraseña cumpliendo los requisitos de seguridad.\n3. Confirmar la nueva contraseña y presionar Guardar Cambios.\n4. Validar en base de datos la estructura del hash generado ($argon2id$v=19$m=19456...).",
     "El backend verifica la clave anterior con Argon2id, genera el nuevo hash con salt criptografico OsRng y actualiza el registro en la base de datos.",
     "Hash actualizado exitosamente en 14 ms. La nueva clave permite iniciar sesion correctamente y la anterior queda revocada.",
     "Critica", "Criptografia", "Aprobado", "Pasa", "Alberto Velazquez", "2h", "diag_07_profile.png"),
     
    ("TC-PRF-004", "Perfil de Usuario", "Ricardo Mendiola", "Rechazo de cambio de contraseña cuando la clave actual es incorrecta", "IRL-IAM-US-05",
     "Dado que el usuario ingresa una contraseña actual erronea\nCuando intenta cambiar la contraseña\nEntonces el servidor rechaza con HTTP 400 Bad Request sin alterar el hash en BD",
     "Usuario en formulario de cambio de clave.",
     "1. Ingresar una contraseña actual incorrecta.\n2. Ingresar nueva contraseña valida y confirmar.\n3. Presionar Guardar Cambios.\n4. Medir respuesta y verificar mensaje de alerta.",
     "El backend detecta la no coincidencia del hash Argon2id y rechaza la peticion con mensaje: La contraseña actual es incorrecta.",
     "Rechazo seguro verificado en 11 ms con codigo 400 Bad Request. El hash en base de datos no fue modificado.",
     "Alta", "Seguridad IAM", "Aprobado", "Pasa", "Luis Zuniga", "1.5h", "s2_01_profile_dialog.png"),

    # MÓDULO 5: WORKSPACE REACTIVO & RUNNER MULTIPLATAFORMA
    ("TC-UX-002", "Workspace Reactivo", "Equipo InnovaSoft", "Navegacion reactiva por pestañas [Chat | Subgrupos | Reuniones]", "General",
     "Dado que el usuario esta dentro de un nodo\nCuando alterna entre las pestañas [Chat], [Subgrupos] y [Reuniones]\nEntonces la vista cambia de forma instantanea sin recargas ni parpadeos",
     "Aplicacion nativa de escritorio en ejecucion dentro de un nodo.",
     "1. Hacer clic sobre la pestaña [💬 Chat].\n2. Cambiar a la pestaña [👥 Subgrupos].\n3. Cambiar a la pestaña [📅 Reuniones].\n4. Evaluar tiempos de transicion y ausencia de parpadeos.",
     "La vista alterna de forma instantanea en menos de 16 ms aprovechando la gestion de estado de Riverpod y la aceleracion por GPU.",
     "Navegacion fluida y reactiva al 100% en todas las pestañas.",
     "Alta", "UX Desktop", "Aprobado", "Pasa", "InnovaSoft", "3h", "s2_06_chat_sprint2_integrated.png"),
     
    ("TC-MAC-001", "macOS Runner", "Alberto Velazquez", "Ejecucion nativa de pruebas de widgets en macOS (darwin-arm64)", "Arquitectura",
     "Dado el entorno macOS desktop darwin-arm64\nCuando se ejecutan las pruebas de widgets de Subgrupos y Reuniones con flutter test\nEntonces pasan al 100% sin excepciones de renderizado",
     "Entorno de desarrollo macOS con Flutter SDK 3.11+ y runner nativo de Darwin configurado.",
     "1. Abrir terminal en el directorio del frontend.\n2. Ejecutar flutter test sobre el entorno macOS desktop darwin-arm64.\n3. Validar smoke test, modelos de perfil de usuario y dialogos de subgrupos y reuniones.",
     "Todos los tests compilan y pasan al 100% mostrando +4: All tests passed!.",
     "4 de 4 pruebas aprobadas en 2.3 segundos en macOS Darwin-arm64 sin errores ni warnings.",
     "Alta", "Compilacion & Tests", "Aprobado", "Pasa", "Alberto Velazquez", "2h", "terminal_flutter_test_sprint2.png"),

    ("TC-API-001", "Backend Suite", "Luis Rivera", "Ejecucion de suite de integracion fullstack y endpoints REST en Rust", "Arquitectura",
     "Dado el servidor backend Actix-web levantado en http://127.0.0.1:8080\nCuando se ejecuta la suite automatizada de endpoints de Sprint 2\nEntonces todos los endpoints responden 200 OK / 201 Created con latencia < 20ms",
     "Base de datos PostgreSQL activa con migraciones 001 y 002 aplicadas.",
     "1. Iniciar servidor backend en Rust mediante cargo run.\n2. Ejecutar suite de pruebas de integracion test_sprint2_fullstack.py.\n3. Verificar codigos HTTP, estructura JSON y tiempos de respuesta.",
     "Todos los endpoints de autenticacion, nodos, mensajes, subgrupos y reuniones responden satisfactoriamente en tiempo record.",
     "Suite de integracion completada al 100% con latencia media de 8.2ms y 0 errores.",
     "Critica", "Integracion REST", "Aprobado", "Pasa", "Luis Rivera", "2.5h", "terminal_backend_sprint2.png")
]

print(f"Total test cases configured: {len(ALL_TEST_CASES_S2)}")

# ═════════════════════════════════════════════════════════════════════════════
# 1. GENERAR Product_Backlog_Nueva_Plantilla_IRONLINK.xlsx
# ═════════════════════════════════════════════════════════════════════════════

def create_product_backlog_excel():
    wb = openpyxl.Workbook()
    ws_portada = wb.active
    ws_portada.title = "Portada"
    
    header_fill = PatternFill(start_color="0B132B", end_color="0B132B", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="00E5FF")
    bold_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    
    # ── SHEET 1: Portada ──
    portada_data = [
        ["UNIVERSIDAD GERARDO BARRIOS", ""],
        ["FACULTAD DE CIENCIA Y TECNOLOGÍA", ""],
        ["INGENIERÍA DE SOFTWARE II — CICLO II-2026", ""],
        ["", ""],
        ["PRODUCT & SPRINT BACKLOG OFICIAL", ""],
        ["SISTEMA ENTERPRISE IRONLINK — SPRINT 2", ""],
        ["", ""],
        ["Proyecto", "IronLink (Desktop & Real-Time Collaboration)"],
        ["Equipo", "Equipo InnovaSoft (Equipo 5)"],
        ["Sprint", "Sprint 2 (Colaboración, Chat, Subgrupos y Reuniones)"],
        ["Docente", "Sandra Beatriz Zúñiga"],
        ["Scrum Master / Architecture Lead", "Ludwin Saúl Vásquez Romero"],
        ["QA Lead / Database & Security", "Luis Alexander Rivera Alvarez"],
        ["Frontend Lead / UI & Tester", "Alberto José Velázquez Paz"],
        ["Backend Dev / API & Tester", "Luis Ángel Zúñiga Menjívar"],
        ["Dev / RTC & Chat Lead", "Ricardo Alberto Mendiola Hernández"],
        ["Dev / Reuniones & Síncrono", "Víctor Arnoldo Iglesias Sandoval"],
        ["Dev / Subgrupos & Workspaces", "José Luis Fuentes Ochoa"],
        ["Fecha Inicio", "10 de agosto 2026"],
        ["Fecha Fin", "24 de agosto 2026"],
        ["Versión", "release-sprint2 (beta-2.0)"],
        ["Estado", "Cerrado / 100% DONE"]
    ]
    
    for row in portada_data:
        ws_portada.append(row)
        
    ws_portada.column_dimensions["A"].width = 36
    ws_portada.column_dimensions["B"].width = 55
    
    for r in range(1, 7):
        ws_portada.cell(r, 1).font = title_font if r in (1, 5) else bold_font
        ws_portada.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
        
    for r in range(8, len(portada_data) + 1):
        c1 = ws_portada.cell(r, 1)
        c2 = ws_portada.cell(r, 2)
        c1.font = bold_font
        c2.font = normal_font
        c1.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    # ── SHEET 2: Product Backlog ──
    ws_pbl = wb.create_sheet(title="Product Backlog")
    pbl_headers = ["ID", "ÉPICA", "Historia de Usuario", "Prioridad (por tamaño)", "Prioridad (numérica)", "Responsables Asignados", "Estimación (Horas)", "Criterios de Aceptación (Gherkin)", "Sprint", "Estado"]
    ws_pbl.append(pbl_headers)
    
    pbl_rows = [
        # SPRINT 1 (DONE)
        ("IRL-IAM-US-01", "Registro e incorporación", "Como usuario nuevo, quiero registrarme con mi nombre y correo, para acceder a la plataforma de forma segura.", "GRANDE", 1, "Ludwin Saul Vasquez Romero", "24 h", "Given que el usuario ingresa datos válidos; When envía el formulario; Then crea cuenta en estado PENDING y hashea con Argon2id.", 1, "DONE"),
        ("IRL-IAM-US-02", "Registro e incorporación", "Como usuario registrado, quiero recibir un correo con código OTP y enlace, para confirmar mi cuenta.", "MEDIANA", 3, "Marielena Velasquez Escobar", "16 h", "Given un usuario PENDING; When ingresa el OTP de 6 dígitos; Then activa la cuenta a estado ACTIVE.", 1, "DONE"),
        ("IRL-IAM-US-04", "Autenticación JWT", "Como usuario con cuenta activa, quiero iniciar sesión con correo y contraseña, para acceder a mis salas.", "MEDIANA", 3, "Ludwin Saul Vasquez Romero", "20 h", "Given credenciales válidas; When hace POST /login; Then emite Access Token JWT (15 min) y Refresh Token (7 días).", 1, "DONE"),
        ("IRL-IAM-US-06", "Gestión de roles RBAC", "Como administrador, quiero asignar roles (Owner/Admin/Member), para controlar permisos de acceso.", "GRANDE", 1, "Luis Alexander Rivera Alvarez", "24 h", "Given un moderador de sala; When asigna roles; Then el sistema valida permisos en middleware y restringe accesos.", 1, "DONE"),
        ("IRL-WKS-US-01", "Workspaces y Nodos", "Como moderador, quiero crear una sala y generar un enlace de acceso cerrado, para que miembros autorizados se unan.", "GRANDE", 1, "Walter Jose Ramirez Perez", "28 h", "Given moderador autenticado; When crea un nodo; Then genera token único de 32 hex y asigna rol OWNER.", 1, "DONE"),
        
        # SPRINT 2 (DONE - INNOVASOFT 104 HORAS)
        ("IRL-WKS-US-03", "Nodos y colaboración", "Como usuario miembro, quiero un chat persistente dentro de cada nodo, para comunicarme con otros miembros fuera de las reuniones en vivo.", "Grande / Must (1)", 1, "Ricardo Mendiola; Alberto Velázquez (QA); Luis Zúñiga (Tester)", "28 h", "Given usuario miembro del nodo; When envía mensaje en canal; Then se inserta en PostgreSQL mensajes y se renderiza en Flutter con avatar y rol.", 2, "DONE"),
        ("IRL-WKS-US-02", "Nodos y colaboración", "Como moderador, quiero gestionar subgrupos dentro de mi nodo, para organizar temas o proyectos con acceso controlado.", "Mediana / Must (3)", 3, "José Fuentes; Alberto Velázquez (QA); Luis Zúñiga (Tester)", "28 h", "Given usuario en nodo; When crea subgrupo público/privado; Then inserta en subgrupos, auto-asigna al creador y gestiona membresías Join/Leave.", 2, "DONE"),
        ("IRL-WKS-US-04", "Calendario y reuniones", "Como moderador, quiero programar reuniones en el calendario del nodo, para que los miembros vean los eventos con anticipación.", "Mediana / Should (3)", 3, "Víctor Iglesias; Alberto Velázquez (QA); Luis Zúñiga (Tester)", "28 h", "Given moderador del nodo; When programa reunión con fecha UTC, duración y enlace Meet; Then se agenda y visualiza con badge ● Programada.", 2, "DONE"),
        ("IRL-IAM-US-05", "Identidad y perfil", "Como usuario, quiero personalizar mi perfil, para identificarme fácilmente en el chat y la lista de miembros.", "Pequeña / Should (5)", 5, "Ricardo Mendiola; Alberto Velázquez (QA); Luis Zúñiga (Tester)", "20 h", "Given usuario autenticado; When selecciona color de avatar (8 opciones), estado de presencia y clave; Then actualiza vía PUT /users/me con Argon2id.", 2, "DONE"),
        
        # SPRINT 3 (FUTURO BACKLOG)
        ("IRL-NTF-US-01", "Notificaciones y alertas", "Como participante, quiero recibir recordatorios cuando una reunión esté por comenzar, para no perderla.", "MEDIANA", 3, "Víctor Iglesias", "20 h", "Given 15 min previos a reunión; When evalúa eventos; Then emite alerta al escritorio.", 3, "BACKLOG"),
        ("IRL-NTF-US-02", "Avisos en tiempo real", "Como participante, quiero notificación en la app cuando el moderador inicie la sesión.", "MEDIANA", 3, "Luis Zúñiga", "20 h", "Given inicio de sala síncrona; When abre sesión; Then notifica a inscritos.", 3, "BACKLOG"),
        ("IRL-NTF-US-03", "Panel de notificaciones", "Como usuario, quiero un panel consolidado de notificaciones agrupadas.", "PEQUEÑA", 5, "Alberto Velázquez", "12 h", "Given avisos acumulados; When abre panel; Then visualiza lista con marcar leídas.", 3, "BACKLOG")
    ]
    
    for r in pbl_rows:
        ws_pbl.append(list(r))
        
    for c in range(1, len(pbl_headers) + 1):
        cell = ws_pbl.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    ws_pbl.column_dimensions["A"].width = 16
    ws_pbl.column_dimensions["B"].width = 24
    ws_pbl.column_dimensions["C"].width = 45
    ws_pbl.column_dimensions["D"].width = 20
    ws_pbl.column_dimensions["E"].width = 18
    ws_pbl.column_dimensions["F"].width = 36
    ws_pbl.column_dimensions["G"].width = 18
    ws_pbl.column_dimensions["H"].width = 45
    ws_pbl.column_dimensions["I"].width = 10
    ws_pbl.column_dimensions["J"].width = 12

    # ── SHEET 3: Sprint Backlog (104 Horas) ──
    ws_sbl = wb.create_sheet(title="Sprint Backlog")
    sbl_headers = ["ID", "ÉPICA", "Historia de Usuario", "Prioridad (por tamaño)", "Prioridad (numérica)", "Sprint", "Dueño de la tarea", "Estimación de esfuerzo", "Estado"]
    base_date = datetime.date(2026, 8, 3)
    date_cols = [str(base_date + datetime.timedelta(days=i)) for i in range(28)]
    ws_sbl.append(sbl_headers + date_cols)
    
    sbl_rows = [
        ("IRL-WKS-US-03", "Nodos y colaboración", "Como usuario miembro, quiero un chat persistente dentro de cada nodo, para comunicarme con otros miembros fuera de las reuniones en vivo.", "Grande / Must (1)", 1, 2, "Ricardo Mendiola; Alberto Velazquez (QA); Luis Zuniga (Tester)", "28 h", "Done",
         [2, 2, 2, 2, 2, 0, 0, 2.5, 2.5, 2.5, 2.5, 2.5, 1, 0, 1.5, 1.5, 1.5, 1.5, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
        ("IRL-WKS-US-02", "Nodos y colaboración", "Como moderador, quiero gestionar subgrupos dentro de mi nodo, para organizar temas o proyectos con acceso controlado.", "Mediana / Must (3)", 3, 2, "Jose Fuentes; Alberto Velazquez (QA); Luis Zuniga (Tester)", "28 h", "Done",
         [0, 0, 0, 0, 0, 0, 0, 2, 2, 2, 2, 2, 1, 1, 2.5, 2.5, 2.5, 2.5, 2, 1, 0, 1, 1, 1, 1, 0, 0, 0]),
        ("IRL-WKS-US-04", "Calendario y reuniones", "Como moderador, quiero programar reuniones en el calendario del nodo, para que los miembros vean los eventos con anticipación.", "Mediana / Should (3)", 3, 2, "Victor Iglesias; Alberto Velazquez (QA); Luis Zuniga (Tester)", "28 h", "Done",
         [1.5, 1.5, 1.5, 1.5, 1, 0, 0, 2, 2, 2, 2, 2, 0, 0, 2, 2, 2, 2, 1.5, 0.5, 0, 0.5, 0.5, 0.5, 0.5, 0, 0, 0]),
        ("IRL-IAM-US-05", "Identidad y perfil", "Como usuario, quiero personalizar mi perfil, para identificarme fácilmente en el chat y la lista de miembros.", "Pequeña / Should (5)", 5, 2, "Ricardo Mendiola; Alberto Velazquez (QA); Luis Zuniga (Tester)", "20 h", "Done",
         [1.5, 1.5, 1.5, 1.5, 1, 0, 0, 1.5, 1.5, 1.5, 1.5, 1, 0, 0, 1, 1, 1, 1, 1, 0, 0, 0.5, 0.5, 0.5, 0.5, 0, 0, 0])
    ]
    for row_info in sbl_rows:
        id_h, ep, hu, pt, pn, sp, own, est, est_s, hours = row_info
        ws_sbl.append([id_h, ep, hu, pt, pn, sp, own, est, est_s] + hours)
        
    for c in range(1, len(sbl_headers) + len(date_cols) + 1):
        cell = ws_sbl.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_sbl.column_dimensions["A"].width = 16
    ws_sbl.column_dimensions["B"].width = 24
    ws_sbl.column_dimensions["C"].width = 45
    ws_sbl.column_dimensions["G"].width = 35

    # ── SHEET 4: BundowChart ──
    ws_bd = wb.create_sheet(title="BundowChart")
    ws_bd.append(["Historia de Sprint 2 (InnovaSoft)", "Est. Inicial", "Sem1", "Sem2", "Sem3", "Sem4", "Total Real"])
    bd_data = [
        ["IRL-WKS-US-03 Chat Persistente", 28, 10, 13.5, 4.5, 0, 28],
        ["IRL-WKS-US-02 Gestion Subgrupos", 28, 0, 11, 11.5, 5.5, 28],
        ["IRL-WKS-US-04 Calendario Reuniones", 28, 7, 8, 10, 3, 28],
        ["IRL-IAM-US-05 Perfil y Presencia", 20, 7, 6, 5, 2, 20],
        ["TOTAL QUEMADO EN LA SEMANA", 104, 24, 38.5, 31, 10.5, 104],
        ["", "", "", "", "", "", ""],
        ["Ajustes de Seguimiento", "Inicio", "Sem1", "Sem2", "Sem3", "Sem4", ""],
        ["Horas Planificadas (Ideal)", 104, 78, 52, 26, 0, ""],
        ["Horas Restantes Reales", 104, 78, 46, 18, 0, ""]
    ]
    for r in bd_data:
        ws_bd.append(r)
        
    for c in range(1, 8):
        cell = ws_bd.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    chart = LineChart()
    chart.title = "Burndown Chart — Sprint 2 IronLink (104 Horas)"
    chart.style = 13
    chart.y_axis.title = "Horas Restantes"
    chart.x_axis.title = "Semanas de Iteración"
    data = Reference(ws_bd, min_col=2, min_row=8, max_col=6, max_row=9)
    cats = Reference(ws_bd, min_col=2, min_row=7, max_col=6, max_row=7)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    ws_bd.add_chart(chart, "A12")

    # ── SHEET 5: Acuerdo QA (DoD y DoR) ──
    ws_ac = wb.create_sheet(title="Acuerdo QA")
    ws_ac.append(["Categoría", "Criterio de Aceptación (Definition of Done - DoD)", "Checklist", "Área", "Estado"])
    dod_data = [
        ["Código", "Sigue los estándares de nomenclatura definidos por el equipo (Rust snake_case, Dart camelCase)", "True", "Código", "True"],
        ["Código", "El código está documentado, tipado y comentado en controladores y servicios", "True", "Gestión Scrum", "True"],
        ["Código", "Fue subido y mergeado correctamente al repositorio GitHub sin conflictos en main", "True", "Pruebas", "True"],
        ["Código", "No presenta errores de compilación, warnings bloqueantes ni fugas de memoria", "True", "Funcionalidad", "True"],
        ["Gestión Scrum", "Las tarjetas de Historias de Usuario fueron actualizadas en el tablero Trello", "True", "Revisión", "True"],
        ["Gestión Scrum", "Se registró el avance y esfuerzo real invertido (104 Horas exactas)", "True", "Base de Datos", "True"],
        ["Gestión Scrum", "La evidencia visual y técnica (capturas/logs) fue adjuntada a cada tarjeta", "True", "Seguridad", "True"],
        ["Gestión Scrum", "La totalidad de las historias del Sprint 2 fueron movidas a la columna [DONE]", "True", "Rendimiento", "True"],
        ["Funcionalidad", "Cumple al 100% con los criterios de aceptación Gherkin definidos en el Product Backlog", "True", "", ""],
        ["Funcionalidad", "Persistencia relacional en PostgreSQL verificada (mensajes, subgrupos, reuniones, users)", "True", "", ""],
        ["Funcionalidad", "Integridad referencial y cascada (ON DELETE CASCADE) probadas en borrados", "True", "", ""],
        ["Funcionalidad", "La interfaz de escritorio en macOS responde de forma reactiva y sin parpadeos", "True", "", ""],
        ["Pruebas QA", "23 Casos de Prueba diseñados y ejecutados con resultado 100% Aprobado / Pasa", "True", "", ""],
        ["Pruebas QA", "Pruebas unitarias de widgets en macOS pasando satisfactoriamente con flutter test", "True", "", ""],
        ["Pruebas QA", "Suite de integración de backend en Rust validada con latencia media < 10ms", "True", "", ""],
        ["Pruebas QA", "5 Bugs detectados durante el ciclo de QA fueron solucionados y cerrados al 100%", "True", "", ""],
        ["Revisión", "Código y arquitectura auditados y aprobados por el QA Lead (Luis Rivera) y Scrum Master (Ludwin Vásquez)", "True", "", ""],
        ["Revisión", "Criptografía Argon2id y seguridad RBAC verificadas contra intrusiones no autorizadas", "True", "", ""]
    ]
    for r in dod_data:
        ws_ac.append(r)
        
    for c in range(1, 6):
        cell = ws_ac.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws_ac.append([])
    ws_ac.append(["Sprint", "Título de la Tarjeta en Trello", "Descripción de Historia", "Checklist 1 — Definition of Ready (DoR) Auditado"])
    
    dor_cards = [
        ("2", "IRL-WKS-US-03: Chat Persistente", "Como usuario miembro, quiero un chat persistente en cada nodo, para comunicarme fuera de reuniones.",
         "☑ Formato estándar (Como/Quiero/Para) validado por PO.\n☑ Criterios Gherkin definidos.\n☑ Estimación acordada: 28h.\n☑ Responsables asignados: Ricardo M., Alberto V., Luis Z.\n☑ Migración SQL 002 lista.\n☑ UX validado en Flutter."),
        ("2", "IRL-WKS-US-02: Gestión de Subgrupos", "Como moderador, quiero gestionar subgrupos dentro de mi nodo, para organizar células de trabajo.",
         "☑ Formato estándar validado por PO.\n☑ Criterios Gherkin definidos.\n☑ Estimación acordada: 28h.\n☑ Responsables asignados: José F., Alberto V., Luis Z.\n☑ Modelo de datos subgrupos y miembros resuelto.\n☑ UI de subgrupos aprobada."),
        ("2", "IRL-WKS-US-04: Calendario y Reuniones", "Como moderador, quiero programar reuniones en el calendario del nodo, para anticipar eventos.",
         "☑ Formato estándar validado por PO.\n☑ Criterios Gherkin definidos.\n☑ Estimación acordada: 28h.\n☑ Responsables asignados: Víctor I., Alberto V., Luis Z.\n☑ Timestamps UTC y Meet integrados.\n☑ Interfaz de agenda validada."),
        ("2", "IRL-IAM-US-05: Identidad y Perfil", "Como usuario, quiero personalizar mi perfil, para identificarme en chat y miembros.",
         "☑ Formato estándar validado por PO.\n☑ Criterios Gherkin definidos.\n☑ Estimación acordada: 20h.\n☑ Responsables asignados: Ricardo M., Alberto V., Luis Z.\n☑ Selector de avatar y presencia definidos.\n☑ Argon2id para password confirmado.")
    ]
    for card in dor_cards:
        ws_ac.append(list(card))
        
    for col_letter, col_w in [("A", 20), ("B", 35), ("C", 45), ("D", 50), ("E", 15)]:
        ws_ac.column_dimensions[col_letter].width = col_w

    file_path = os.path.join(OUTPUT_S2_DIR, "Product_Backlog_Nueva_Plantilla_IRONLINK.xlsx")
    wb.save(file_path)
    print(f"✅ Product Backlog Sprint 2 generado en: {file_path}")

create_product_backlog_excel()

# ═════════════════════════════════════════════════════════════════════════════
# 2. GENERAR IronLink_QA_Plan_Sprint2.xlsx (23 CASOS + MATRIZ)
# ═════════════════════════════════════════════════════════════════════════════

def create_qa_plan_excel():
    wb = openpyxl.Workbook()
    ws_portada = wb.active
    ws_portada.title = "Portada"
    
    header_fill = PatternFill(start_color="0B132B", end_color="0B132B", fill_type="solid")
    sub_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    label_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=13, bold=True, color="00E5FF")
    bold_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=9.5)
    pass_font = Font(name="Arial", size=11, bold=True, color="16A34A")
    
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # ── SHEET 1: Portada QA Plan ──
    portada_rows = [
        ["UNIVERSIDAD GERARDO BARRIOS", ""],
        ["FACULTAD DE CIENCIA Y TECNOLOGÍA", ""],
        ["INGENIERÍA DE SOFTWARE II", ""],
        ["", ""],
        ["PLAN DE ASEGURAMIENTO DE LA CALIDAD (QA)", ""],
        ["SUITE COMPLETA DE PRUEBAS DE ESCRITORIO & INTEGRACIÓN — SPRINT 2", ""],
        ["", ""],
        ["Proyecto", "IronLink (Desktop & Real-Time Collaboration)"],
        ["Equipo", "InnovaSoft (Equipo 5 — 7 Integrantes)"],
        ["Sprint", "Sprint 2 (Colaboración, Chat, Subgrupos y Reuniones)"],
        ["Docente", "Sandra Beatriz Zúñiga"],
        ["Scrum Master / Architecture Lead", "Ludwin Saúl Vásquez Romero"],
        ["QA Lead / Database & Security", "Luis Alexander Rivera Alvarez"],
        ["Frontend Lead / Desktop UI & Tester", "Alberto José Velázquez Paz"],
        ["Backend Dev / API & Tester", "Luis Ángel Zúñiga Menjívar"],
        ["Dev / RTC & Chat Lead", "Ricardo Alberto Mendiola Hernández"],
        ["Dev / Reuniones & Síncrono", "Víctor Arnoldo Iglesias Sandoval"],
        ["Dev / Subgrupos & Workspaces", "José Luis Fuentes Ochoa"],
        ["Fecha de Planificación", "10 de agosto 2026"],
        ["Fecha de Ejecución y Cierre", "24 de agosto 2026"],
        ["Total de Casos de Prueba", "23 Casos Diseñados y Ejecutados"],
        ["Resultado Global", "23 / 23 APROBADOS (100% Exitoso)"],
        ["Versión", "release-sprint2 (beta-2.0)"],
        ["Estado", "Aprobado / Certificado para Producción"]
    ]
    for row in portada_rows:
        ws_portada.append(row)
        
    ws_portada.column_dimensions["A"].width = 38
    ws_portada.column_dimensions["B"].width = 58
    
    for r in range(1, 7):
        ws_portada.cell(r, 1).font = title_font if r in (1, 5) else bold_font
        
    for r in range(8, len(portada_rows) + 1):
        c1 = ws_portada.cell(r, 1)
        c2 = ws_portada.cell(r, 2)
        c1.font = bold_font
        c2.font = normal_font
        c1.fill = label_fill

    # ── SHEETS 2 A 24: 23 HOJAS DE CASOS DE PRUEBA ──
    for tc in ALL_TEST_CASES_S2:
        tc_id, module, author, title, hu, gherkin, precond, steps, expected, obtained, priority, test_type, design_st, exec_st, qa_resp, time_est, img_name = tc
        
        ws = wb.create_sheet(title=tc_id)
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 38
        
        # R1: Title banner
        ws.merge_cells("A1:D1")
        cell_t = ws.cell(1, 1, f"CASO DE PRUEBA – {tc_id}  |  {title.upper()}")
        cell_t.fill = header_fill
        cell_t.font = title_font
        cell_t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Metadata Table
        meta_fields = [
            [("ID Caso de Prueba:", tc_id), ("Historia de Usuario:", hu)],
            [("Módulo:", module), ("Tipo de Prueba:", test_type)],
            [("Prioridad:", priority), ("Elaborado por:", author)],
            [("Responsable QA:", qa_resp), ("Estado Diseño:", design_st)],
            [("Fecha Creación:", "10/08/2026"), ("Fecha Ejecución:", "24/08/2026")]
        ]
        
        cur_r = 2
        for pair in meta_fields:
            ws.cell(cur_r, 1, pair[0][0]).font = bold_font
            ws.cell(cur_r, 1).fill = label_fill
            ws.cell(cur_r, 2, pair[0][1]).font = normal_font
            ws.cell(cur_r, 3, pair[1][0]).font = bold_font
            ws.cell(cur_r, 3).fill = label_fill
            ws.cell(cur_r, 4, pair[1][1]).font = normal_font
            for c in range(1, 5):
                ws.cell(cur_r, c).border = thin_border
            cur_r += 1
            
        cur_r += 1 # Empty row
        
        # Section blocks
        sections = [
            ("CRITERIOS DE ACEPTACIÓN (GHERKIN)", gherkin),
            ("PRECONDICIONES Y ENTORNO", precond),
            ("PASOS DETALLADOS DE EJECUCIÓN", steps),
            ("RESULTADO ESPERADO", expected),
            ("RESULTADO OBTENIDO", obtained)
        ]
        
        for sec_title, sec_content in sections:
            ws.merge_cells(start_row=cur_r, start_column=1, end_row=cur_r, end_column=4)
            c_hdr = ws.cell(cur_r, 1, sec_title)
            c_hdr.fill = sub_fill
            c_hdr.font = header_font
            c_hdr.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[cur_r].height = 20
            cur_r += 1
            
            ws.merge_cells(start_row=cur_r, start_column=1, end_row=cur_r, end_column=4)
            c_txt = ws.cell(cur_r, 1, sec_content)
            c_txt.font = normal_font
            c_txt.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for c in range(1, 5):
                ws.cell(cur_r, c).border = thin_border
            ws.row_dimensions[cur_r].height = max(35, len(sec_content.split("\n")) * 16)
            cur_r += 2
            
        # Execution Footer
        ws.cell(cur_r, 1, "ESTADO DE EJECUCIÓN:").font = bold_font
        ws.cell(cur_r, 1).fill = label_fill
        c_pasa = ws.cell(cur_r, 2, f"✔ {exec_st.upper()} (100%)")
        c_pasa.font = pass_font
        c_pasa.fill = pass_fill
        c_pasa.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(cur_r, 3, "TIEMPO REAL INVERTIDO:").font = bold_font
        ws.cell(cur_r, 3).fill = label_fill
        ws.cell(cur_r, 4, time_est).font = bold_font
        
        for c in range(1, 5):
            ws.cell(cur_r, c).border = thin_border
        ws.row_dimensions[cur_r].height = 25

    # ── SHEET 25: Matriz de Trazabilidad ──
    ws_mtx = wb.create_sheet(title="Matriz de Trazabilidad")
    mtx_headers = ["HU ID", "Escenario / Funcionalidad", "Caso de Prueba", "Módulo", "Tipo de Prueba", "Prioridad", "Responsable QA", "Estado Diseño", "Estado Ejecución", "Bugs Detectados / Notas"]
    ws_mtx.append(mtx_headers)
    
    for c in range(1, len(mtx_headers) + 1):
        cell = ws_mtx.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_mtx.row_dimensions[1].height = 28
    
    bug_map = {
        "TC-CHT-001": "BUG-S2-001 (Solucionado - Serialización JSON timestamp UTC)",
        "TC-CHT-004": "BUG-S2-004 (Solucionado - Trim de espacios en blanco en chat)",
        "TC-SUB-001": "BUG-S2-002 (Solucionado - Auto-asignación de membresía en subgrupos)",
        "TC-SUB-006": "BUG-S2-005 (Solucionado - Integridad referencial en borrado en cascada)",
        "TC-REU-001": "BUG-S2-003 (Solucionado - Formateo de fecha y badges en calendario)",
        "TC-PRF-001": "Verificado al 100% (Sin incidencias)",
        "TC-PRF-003": "Verificado al 100% (Argon2id robusto)"
    }
    
    for r_idx, tc in enumerate(ALL_TEST_CASES_S2, start=2):
        tc_id, module, author, title, hu, gherkin, precond, steps, expected, obtained, priority, test_type, design_st, exec_st, qa_resp, time_est, img_name = tc
        notes = bug_map.get(tc_id, "Pasa sin defectos bloqueantes")
        
        row_vals = [hu, title, tc_id, module, test_type, priority, qa_resp, design_st, exec_st, notes]
        ws_mtx.append(row_vals)
        
        for c in range(1, len(row_vals) + 1):
            cell = ws_mtx.cell(row=r_idx, column=c)
            cell.font = normal_font
            cell.border = thin_border
            if c in (1, 3, 6, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 9:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = pass_font
                cell.fill = pass_fill
                
    ws_mtx.column_dimensions["A"].width = 16
    ws_mtx.column_dimensions["B"].width = 40
    ws_mtx.column_dimensions["C"].width = 16
    ws_mtx.column_dimensions["D"].width = 20
    ws_mtx.column_dimensions["E"].width = 18
    ws_mtx.column_dimensions["F"].width = 12
    ws_mtx.column_dimensions["G"].width = 24
    ws_mtx.column_dimensions["H"].width = 16
    ws_mtx.column_dimensions["I"].width = 16
    ws_mtx.column_dimensions["J"].width = 45

    file_path = os.path.join(OUTPUT_S2_DIR, "IronLink_QA_Plan_Sprint2.xlsx")
    wb.save(file_path)
    print(f"✅ QA Plan Excel con 23 casos y Matriz de Trazabilidad generado en: {file_path}")

create_qa_plan_excel()
